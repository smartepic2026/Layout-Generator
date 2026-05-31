"""urs_parser — URS xlsx → RuleEngineInput 변환 정식 모듈.

이전 핸드오프(2026-05-28) §6.3 우선순위 1 작업: `_demo_run.py` 안에 박혀있던
미니 파서를 분리해 반복 사용 가능한 모듈로 승격한 것이다. v0.3 보고서 안건
#2 (URS Excel만 사용) 의 후속 정식화.

스키마 가정 (URS_ConceptualDesign for layout_0516.xlsx 기준):
    시트 2 "Room 요구 규격서"
        - row 4 (B~G): 건축물 메타 (floor, area, dimensions, 동선 시계방향)
        - row 8 이후: Room 명세
            컬럼 인덱스 (0-base): 1=구분, 2=한글명, 3=영문명, 4=청정도,
                                  5=ACH, 6=동선, 7=복장, 8=공정No, 9=면적 비율 (%)
    시트 3 "제조 장비 규격서"
        - row 4 이후 장비 명세
            컬럼 인덱스: 3=room_en, 4=instance_id, 5="W*D*H" 문자열,
                          6=weight_kg, 7=max_operating_weight_kg, 8=process_no

Public API:
    URS_PATH                      — 기본 URS 파일 위치 (세션 독립적).
    parse_urs_xlsx(path)          — 시트 2/3 파싱 → (rooms_dict, equipment_dict, building_info).
    build_rule_engine_input(...)  — 파싱 결과를 RuleEngineInput 으로 조립.
    load_urs_as_input(path=...)   — 위 두 단계를 묶은 한 줄 helper.
    clock_from_text(text)         — "12시 방향" → "12" 와 같은 정규화.

설계 원칙:
    - openpyxl 외 외부 의존 없음.
    - 셀 누락에 대해 fallback 을 제공하되, 가급적 명시적 default 값.
    - 룰 엔진 본체에 대한 단일 진입점 — 다른 에이전트가 동일한 input 을
      만들고 싶을 때도 본 모듈을 import 하면 됨.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from .models import (
    BuildingSpec,
    FlowPolicy,
    OrganizationSpec,
    Overrides,
    ProductSpec,
    RuleEngineInput,
)


# ---------------------------------------------------------------------------
# 기본 URS 파일 위치 — 세션 ID 가 바뀌어도 동작하도록 모듈 위치 기준.
# rule_engine/urs_parser.py → 부모 rule_engine/, 그 부모 claude_cowork_file/.
# ---------------------------------------------------------------------------

URS_PATH: Path = (
    Path(__file__).resolve().parent.parent
    / "URS_ConceptualDesign for layout_0516.xlsx"
)


# ---------------------------------------------------------------------------
# 시트/열 위치 상수 — 스키마가 바뀌면 여기만 수정.
# ---------------------------------------------------------------------------

_SHEET_ROOMS = "Room 요구 규격서"
_SHEET_EQUIPMENT = "제조 장비 규격서"

_ROOM_DATA_START_ROW = 8
_EQUIPMENT_DATA_START_ROW = 4
_BUILDING_META_ROW = 4

# Room 컬럼 (0-base index, openpyxl iter_rows values_only=True 기준).
_ROOM_COL_CATEGORY = 1
_ROOM_COL_NAME_KO = 2
_ROOM_COL_NAME_EN = 3
_ROOM_COL_CLEAN_GRADE = 4
_ROOM_COL_ACPH = 5
_ROOM_COL_ROOM_FLOW = 6
_ROOM_COL_GOWNING = 7
_ROOM_COL_PROCESS_NO = 8
_ROOM_COL_AREA_RATIO_PCT = 9  # 회의 안건 #5 (2026-05-26)

# Equipment 컬럼.
_EQ_COL_ROOM_EN = 3
_EQ_COL_INSTANCE_ID = 4
_EQ_COL_DIM_STRING = 5  # "W*D*H"
_EQ_COL_WEIGHT = 6
_EQ_COL_MAX_OP_WEIGHT = 7
_EQ_COL_PROCESS_NO = 8


# ---------------------------------------------------------------------------
# 문자열 정규화 헬퍼
# ---------------------------------------------------------------------------

def clock_from_text(text: Any) -> str:
    """'12시 방향', '3시', '9 o\\'clock' → '12' / '3' / '9'.

    URS 에 들어오는 자유 형식 문자열을 ClockDirection enum 값으로 정규화.
    유효 시계 방향이 추출되지 않으면 "12" fallback.
    """
    if not text:
        return "12"
    digits = "".join(c for c in str(text) if c.isdigit())
    return digits if digits in ("12", "3", "6", "9") else "12"


def _coerce_float(value: Any) -> float | None:
    """openpyxl 셀 값을 float 로 변환. 실패하면 None."""
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.strip().lower() == "not applicable":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# 시트 2 — Room 명세
# ---------------------------------------------------------------------------

def _parse_building_meta(ws_room) -> dict:
    """시트 2 의 row 4 에서 건축물 메타 6 컬럼 추출."""
    r = _BUILDING_META_ROW
    return {
        "floor_level_str": ws_room.cell(row=r, column=2).value,
        "total_area_m2": ws_room.cell(row=r, column=3).value,
        "dimensions": ws_room.cell(row=r, column=4).value,
        "personnel": ws_room.cell(row=r, column=5).value,
        "material_in": ws_room.cell(row=r, column=6).value,
        "waste_out": ws_room.cell(row=r, column=7).value,
    }


def _clean_name(value) -> str:
    """방·장비 이름의 앞뒤 공백 제거 (Doc Agent 매칭 실패 방지, 2026-05-29).

    URS xlsx 원본에 trailing space ('Cell Culture ') 가 섞여 있어
    adjacency 의 이름 기반 매칭이 실패하던 문제를 원천에서 차단한다.
    None 은 그대로 통과.
    """
    if isinstance(value, str):
        return value.strip()
    return value


def _parse_room_row(row: tuple) -> dict | None:
    """한 줄을 Room dict 로 변환. skip 대상이면 None."""
    name_en = _clean_name(row[_ROOM_COL_NAME_EN])
    if not name_en:
        return None
    if isinstance(name_en, str) and "합계" in name_en:
        return None

    clean_grade_raw = row[_ROOM_COL_CLEAN_GRADE]
    clean_grade = (
        clean_grade_raw.replace("Grade ", "").strip()
        if isinstance(clean_grade_raw, str)
        else (clean_grade_raw or "D")
    )

    process_no_str = row[_ROOM_COL_PROCESS_NO] or ""
    process_no = [p.strip() for p in process_no_str.split(",") if p.strip()]

    area_ratio_raw = (
        row[_ROOM_COL_AREA_RATIO_PCT]
        if len(row) > _ROOM_COL_AREA_RATIO_PCT
        else None
    )

    return {
        "name_ko": _clean_name(row[_ROOM_COL_NAME_KO]) or name_en,
        "name_en": name_en,
        "category": row[_ROOM_COL_CATEGORY],
        "clean_grade": clean_grade,
        "room_flow": row[_ROOM_COL_ROOM_FLOW] or "Both-way",
        "gowning_type": row[_ROOM_COL_GOWNING] or "일상복",
        "process_no": process_no,
        "air_changes_per_hour": _coerce_float(row[_ROOM_COL_ACPH]),
        "area_ratio_pct": _coerce_float(area_ratio_raw),
    }


# ---------------------------------------------------------------------------
# 시트 3 — Equipment 명세
# ---------------------------------------------------------------------------

def _parse_equipment_row(row: tuple) -> dict | None:
    """한 줄을 Equipment dict 로 변환. skip 대상이면 None."""
    room_en = row[_EQ_COL_ROOM_EN]
    dim_str = row[_EQ_COL_DIM_STRING]
    if not room_en or not dim_str:
        return None
    try:
        w, d, h = [int(s.strip()) for s in str(dim_str).split("*")]
    except (ValueError, AttributeError):
        return None

    return {
        "room_en": _clean_name(room_en),
        "instance_id": _clean_name(row[_EQ_COL_INSTANCE_ID]),
        "name": _clean_name(row[_EQ_COL_INSTANCE_ID]),
        "W": w,
        "D": d,
        "H": h,
        "weight_kg": row[_EQ_COL_WEIGHT],
        "max_operating_weight_kg": row[_EQ_COL_MAX_OP_WEIGHT],
        "process_no": [
            p.strip()
            for p in (row[_EQ_COL_PROCESS_NO] or "").split(",")
            if p.strip()
        ],
    }


# ---------------------------------------------------------------------------
# 진입점 — Public API
# ---------------------------------------------------------------------------

def parse_urs_xlsx(
    xlsx_path: str | Path,
) -> tuple[list[dict], list[dict], dict]:
    """URS xlsx 의 시트 2/3 를 읽어 dict 형태로 추출.

    Args:
        xlsx_path: URS xlsx 파일 경로.

    Returns:
        (urs_rooms, urs_equipment, building_info) 튜플.
            urs_rooms: list[dict] — Room 한 점당 dict.
            urs_equipment: list[dict] — Equipment 한 점당 dict.
            building_info: dict — 건축물 메타.

    Raises:
        FileNotFoundError: 파일이 없을 때.
        KeyError: 기대 시트명이 없을 때 (openpyxl).
    """
    p = Path(xlsx_path)
    if not p.exists():
        raise FileNotFoundError(f"URS xlsx 파일이 없습니다: {p}")
    wb = openpyxl.load_workbook(p, data_only=True)

    ws_room = wb[_SHEET_ROOMS]
    building_info = _parse_building_meta(ws_room)

    urs_rooms: list[dict] = []
    for row in ws_room.iter_rows(min_row=_ROOM_DATA_START_ROW, values_only=True):
        parsed = _parse_room_row(row)
        if parsed is not None:
            urs_rooms.append(parsed)

    ws_eq = wb[_SHEET_EQUIPMENT]
    urs_equipment: list[dict] = []
    for row in ws_eq.iter_rows(
        min_row=_EQUIPMENT_DATA_START_ROW, values_only=True
    ):
        parsed = _parse_equipment_row(row)
        if parsed is not None:
            urs_equipment.append(parsed)

    return urs_rooms, urs_equipment, building_info


# ---------------------------------------------------------------------------
# 건축물 메타 → BuildingSpec 변환
# ---------------------------------------------------------------------------

_DEFAULT_BUILDING_WIDTH_MM = 78500
_DEFAULT_BUILDING_DEPTH_MM = 42500
_DEFAULT_TOTAL_AREA_M2 = 3340.0


def _parse_dimensions(text: Any) -> tuple[int, int]:
    """'78500mm * 42500mm' → (78500, 42500). 실패 시 default."""
    raw = (text or "").replace("mm", "")
    try:
        w, d = [int(s.strip()) for s in raw.split("*")]
        return w, d
    except Exception:
        return _DEFAULT_BUILDING_WIDTH_MM, _DEFAULT_BUILDING_DEPTH_MM


def _parse_floor_level(text: Any) -> int:
    """'3층' → 3. 실패 시 1."""
    raw = text or "1층"
    try:
        return int("".join(c for c in str(raw) if c.isdigit()))
    except ValueError:
        return 1


def build_rule_engine_input(
    urs_rooms: list[dict],
    urs_equipment: list[dict],
    building_info: dict,
    *,
    product: ProductSpec | None = None,
    flow_policy: FlowPolicy | None = None,
    organization: OrganizationSpec | None = None,
    overrides: Overrides | None = None,
) -> RuleEngineInput:
    """파싱 결과 dict 묶음을 RuleEngineInput 으로 조립.

    Args:
        urs_rooms: parse_urs_xlsx 가 돌려준 rooms dict 리스트.
        urs_equipment: parse_urs_xlsx 가 돌려준 equipment dict 리스트.
        building_info: parse_urs_xlsx 가 돌려준 건축물 메타.
        product: 옵션. None 이면 v1 기본값 (mAb / 8000L).
        flow_policy: 옵션. None 이면 v1 기본값.
        organization: 옵션. None 이면 v1 기본값.
        overrides: 옵션. None 이면 빈 Overrides.

    Returns:
        rule_engine.run_rule_engine 에 그대로 넘길 수 있는 RuleEngineInput.

    Note:
        v1 (mAb DS) 기본값 이외의 modality 가 들어오면 ProductSpec 을 직접
        구성해서 product 인자로 넘겨야 함. 룰 엔진 본체는 mAb 만 지원.
    """
    w, d = _parse_dimensions(building_info.get("dimensions"))
    floor = _parse_floor_level(building_info.get("floor_level_str"))

    return RuleEngineInput(
        product=product or ProductSpec(
            modality="mAb",
            culture_scale_L=8000,
            n_product_types=1,
            virus_filtration_required=True,
            closed_system_main_process=False,
        ),
        building=BuildingSpec(
            total_floor_area_m2=float(
                building_info.get("total_area_m2") or _DEFAULT_TOTAL_AREA_M2
            ),
            width_mm=w,
            depth_mm=d,
            floor_level=floor,
            personnel_entrance=clock_from_text(building_info.get("personnel")),
            material_inlet=clock_from_text(building_info.get("material_in")),
            waste_outlet=clock_from_text(building_info.get("waste_out")),
            elevator_for_material_in=clock_from_text(
                building_info.get("material_in")
            ),
            elevator_for_waste_out=clock_from_text(
                building_info.get("waste_out")
            ),
        ),
        flow_policy=flow_policy or FlowPolicy(
            airlock_default_type="cascade",
            supply_return_corridor_separate=True,
            biological_safety_isolation=False,
        ),
        organization=organization or OrganizationSpec(
            gender_separated_gowning=True,
            include_office_onsite=True,
            include_toilet_onsite=True,
            include_monitoring_room_onsite=True,
            include_lobby_onsite=True,
        ),
        overrides=overrides or Overrides(),
        urs_rooms=urs_rooms,
        urs_equipment=urs_equipment,
    )


def load_urs_as_input(path: str | Path = URS_PATH, **kwargs) -> RuleEngineInput:
    """parse_urs_xlsx → build_rule_engine_input 한 줄 helper.

    Example:
        >>> from rule_engine import load_urs_as_input, run_rule_engine
        >>> output = run_rule_engine(load_urs_as_input())
        >>> output.meta["stats"]["rooms"]
        48
    """
    urs_rooms, urs_equipment, building_info = parse_urs_xlsx(path)
    return build_rule_engine_input(
        urs_rooms, urs_equipment, building_info, **kwargs
    )
